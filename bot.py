import logging
import os
import tempfile
from collections import Counter

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.constants import ChatAction
from telegram.error import TimedOut, NetworkError
from telegram.ext import (
    Application, CommandHandler, MessageHandler, CallbackQueryHandler,
    ContextTypes, filters,
)

from parsers.detect import parse_statement, BANK_LABELS
from parsers import kasir as kasir_parser
from parsers import preformatted as preformatted_parser
from parsers.common import write_xlsx, write_recon_xlsx, build_filename, sheet_title_from_meta, split_workbook
from openpyxl import load_workbook

logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

BOT_TOKEN = os.environ.get('BOT_TOKEN', '')

WELCOME = (
    "Halo! Kirim aku:\n"
    "- PDF rekening koran (BCA, BRI, atau Bank Jago),\n"
    "- XLSX rekap kasir, atau\n"
    "- XLSX rekening yang sudah diolah (format 9 kolom bot ini)\n\n"
    "Nanti aku ubah/rapikan jadi XLSX format seragam: Tanggal, Keterangan Transaksi, "
    "Kategori Transaksi, Debit, Kredit, Saldo Kumulatif, Subjek Transaksi, "
    "Objek Transaksi, Keterangan Tambahan.\n\n"
    "Upload beberapa rekening lalu ketik /gabung untuk menggabungkan semuanya "
    "jadi satu file rekonsiliasi (satu sheet per rekening).\n\n"
    "Ketik /sesi untuk lihat rekening apa saja yang sudah tersimpan, atau "
    "/reset untuk mengosongkan semua sebelum menggabung (hindari salah gabung).\n\n"
    "Kalau upload XLSX yang sheet-nya lebih dari satu (misal hasil /gabung), "
    "otomatis dipecah lagi jadi file terpisah per sheet."
)

# In-memory session per chat: list of {'label', 'rows', 'saldo_awal', 'saldo_akhir'}.
# Lives only as long as the bot process runs — resets on redeploy/restart.
SESSIONS = {}
# Which entry indices are currently checked in the /gabung picker, per chat.
SELECTIONS = {}


def _session_add(chat_id, label, rows, saldo_awal, saldo_akhir, meta=None):
    SESSIONS.setdefault(chat_id, []).append({
        'label': label, 'rows': rows, 'saldo_awal': saldo_awal, 'saldo_akhir': saldo_akhir,
        'bulan': (meta or {}).get('bulan', ''), 'tahun': (meta or {}).get('tahun', ''),
    })


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(WELCOME)


async def handle_pdf(update: Update, context: ContextTypes.DEFAULT_TYPE, doc, tmp):
    pdf_path = os.path.join(tmp, doc.file_name)

    tg_file = await doc.get_file()
    await tg_file.download_to_drive(pdf_path)

    bank, info, xlsx_path, rows, meta = parse_statement(pdf_path, tmp)
    label = sheet_title_from_meta(meta)
    _session_add(update.effective_chat.id, label, rows, info.get('saldo_awal'), info.get('saldo_akhir'), meta)

    caption_lines = [
        f"Bank terdeteksi: {BANK_LABELS.get(bank, bank)}",
        f"Total baris transaksi: {info.get('jumlah_baris')}",
    ]
    if info.get('warnings'):
        caption_lines.append(f"⚠️ {len(info['warnings'])} baris saldo tidak cocok checkpoint, cek manual.")
    if info.get('warning'):
        caption_lines.append(f"⚠️ {info['warning']}")
    caption_lines.append(f"Tersimpan di sesi sebagai \"{label}\" ({len(SESSIONS[update.effective_chat.id])} rekening total). Ketik /gabung untuk menggabungkan.")
    return xlsx_path, caption_lines


async def handle_xlsx(update: Update, context: ContextTypes.DEFAULT_TYPE, doc, tmp, existing_path=None):
    src_path = existing_path
    if src_path is None:
        src_path = os.path.join(tmp, doc.file_name)
        tg_file = await doc.get_file()
        await tg_file.download_to_drive(src_path)

    # An uploaded .xlsx could be a cashier report OR a statement that's
    # already in this bot's own 9-column format (a manually reconstructed
    # old statement, or a previous output the user annotated). Try the
    # latter first since its header is unambiguous.
    if preformatted_parser.is_preformatted(src_path):
        rows, saldo_awal, saldo_akhir, info, meta = preformatted_parser.build_rows(src_path)
        filename = build_filename(meta['self_code'], meta['bulan'], meta['tahun'])
        xlsx_path = os.path.join(tmp, filename)
        write_xlsx(rows, xlsx_path, saldo_awal=saldo_awal, saldo_akhir=saldo_akhir)

        label = sheet_title_from_meta(meta)
        _session_add(update.effective_chat.id, label, rows, saldo_awal, saldo_akhir, meta)

        caption_lines = [
            f"Rekening terdeteksi (format sudah diolah): {meta['self_code']}",
            f"Total baris transaksi: {len(rows)}",
        ]
        if info.get('selisih_flags'):
            caption_lines.append(f"⚠️ {info['selisih_flags']} baris punya selisih != 0 di kolom Selisih, cek manual.")
        caption_lines.append(f"Tersimpan di sesi sebagai \"{label}\" ({len(SESSIONS[update.effective_chat.id])} rekening total). Ketik /gabung untuk menggabungkan.")
        return xlsx_path, caption_lines

    rows, saldo_awal, saldo_akhir, info, meta = kasir_parser.build_rows(src_path)
    filename = build_filename(meta['self_code'], meta['bulan'], meta['tahun'])
    xlsx_path = os.path.join(tmp, filename)
    write_xlsx(rows, xlsx_path, saldo_awal=saldo_awal, saldo_akhir=saldo_akhir)

    label = sheet_title_from_meta(meta)
    _session_add(update.effective_chat.id, label, rows, saldo_awal, saldo_akhir, meta)

    caption_lines = [
        "Kasir Stoa Space (dikonversi ke format seragam)",
        f"Total baris transaksi: {len(rows)}",
        f"Toko dikenali: {info['toko_dikenali']}, Tenant Lain: {info['tenant_lain']}",
        f"Tersimpan di sesi sebagai \"{label}\" ({len(SESSIONS[update.effective_chat.id])} rekening total). Ketik /gabung untuk menggabungkan.",
    ]
    return xlsx_path, caption_lines


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    doc = update.message.document
    if not doc:
        return
    name = doc.file_name.lower()
    is_pdf = name.endswith('.pdf')
    is_xlsx = name.endswith('.xlsx')
    if not (is_pdf or is_xlsx):
        await update.message.reply_text('Kirim file PDF (rekening koran) atau XLSX (rekap kasir) ya.')
        return

    await context.bot.send_chat_action(chat_id=update.effective_chat.id, action=ChatAction.TYPING)
    status_msg = await update.message.reply_text('Lagi diproses...')

    with tempfile.TemporaryDirectory() as tmp:
        try:
            if is_xlsx:
                # kalau file punya lebih dari 1 sheet, anggap ini permintaan
                # "pisahkan per sheet" (kebalikan dari /gabung) — bukan satu
                # statement/kasir tunggal untuk di-parse.
                probe_path = os.path.join(tmp, doc.file_name)
                tg_file = await doc.get_file()
                await tg_file.download_to_drive(probe_path)
                probe_wb = load_workbook(probe_path, read_only=True)
                sheet_count = len(probe_wb.sheetnames)
                probe_wb.close()

                if sheet_count > 1:
                    await handle_split(update, context, probe_path, tmp, status_msg)
                    return
                xlsx_path, caption_lines = await handle_xlsx(update, context, doc, tmp, existing_path=probe_path)
            else:
                xlsx_path, caption_lines = await handle_pdf(update, context, doc, tmp)
        except (TimedOut, NetworkError):
            logger.exception('Timeout jaringan saat memproses %s', doc.file_name)
            await status_msg.edit_text(
                'Koneksi ke Telegram sempat timeout (biasanya sesaat setelah bot baru redeploy). '
                'Coba kirim ulang file-nya.'
            )
            return
        except ValueError as e:
            await status_msg.edit_text(str(e))
            return
        except Exception:
            logger.exception('Gagal memproses %s', doc.file_name)
            await status_msg.edit_text(
                'Gagal memproses file ini. Kemungkinan formatnya sedikit beda dari '
                'yang sudah aku pelajari — kirim ke admin untuk dicek.'
            )
            return

        await status_msg.edit_text('\n'.join(caption_lines))
        with open(xlsx_path, 'rb') as f:
            await update.message.reply_document(document=f, filename=os.path.basename(xlsx_path))


async def handle_split(update: Update, context: ContextTypes.DEFAULT_TYPE, src_path, tmp, status_msg):
    outputs = split_workbook(src_path, tmp)
    await status_msg.edit_text(f'File ini punya {len(outputs)} sheet — dipecah jadi {len(outputs)} file terpisah:')
    for sheet_name, out_path in outputs:
        with open(out_path, 'rb') as f:
            await context.bot.send_document(
                chat_id=update.effective_chat.id, document=f, filename=os.path.basename(out_path),
                caption=sheet_name,
            )


def _gabung_keyboard(chat_id):
    entries = SESSIONS.get(chat_id, [])
    selected = SELECTIONS.setdefault(chat_id, set())
    rows = []
    for i, e in enumerate(entries):
        mark = '✅' if i in selected else '⬜'
        rows.append([
            InlineKeyboardButton(f'{mark} {e["label"]}', callback_data=f'toggle:{i}'),
            InlineKeyboardButton('🗑', callback_data=f'remove:{i}'),
        ])
    rows.append([
        InlineKeyboardButton('✅ Pilih semua', callback_data='select_all'),
        InlineKeyboardButton('⬜ Kosongkan pilihan', callback_data='select_none'),
    ])
    rows.append([
        InlineKeyboardButton('📎 Proses Gabung', callback_data='process'),
        InlineKeyboardButton('🧹 Hapus Semua Sesi', callback_data='reset_all'),
    ])
    rows.append([InlineKeyboardButton('Batal', callback_data='cancel')])
    return InlineKeyboardMarkup(rows)


async def sesi_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    entries = SESSIONS.get(chat_id, [])
    if not entries:
        await update.message.reply_text('Sesi kosong — belum ada rekening yang diproses.')
        return
    lines = [f'{i + 1}. {e["label"]}' for i, e in enumerate(entries)]
    await update.message.reply_text(
        f'Ada {len(entries)} rekening tersimpan di sesi ini:\n' + '\n'.join(lines) +
        '\n\nKetik /gabung untuk menggabungkan, atau /reset untuk mengosongkan semua.'
    )


async def reset_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    count = len(SESSIONS.get(chat_id, []))
    SESSIONS[chat_id] = []
    SELECTIONS[chat_id] = set()
    if count:
        await update.message.reply_text(f'Sesi dikosongkan — {count} rekening yang tersimpan sudah dihapus.')
    else:
        await update.message.reply_text('Sesi memang sudah kosong.')


async def gabung_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    entries = SESSIONS.get(chat_id, [])
    if not entries:
        await update.message.reply_text(
            'Belum ada rekening yang diproses di sesi ini. Upload dulu PDF/XLSX-nya, baru ketik /gabung.'
        )
        return
    SELECTIONS[chat_id] = set(range(len(entries)))  # default: semua terpilih
    await update.message.reply_text(
        'Pilih rekening yang mau digabung jadi satu file rekonsiliasi (satu sheet per rekening):',
        reply_markup=_gabung_keyboard(chat_id),
    )


async def gabung_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    chat_id = update.effective_chat.id
    entries = SESSIONS.get(chat_id, [])
    selected = SELECTIONS.setdefault(chat_id, set())
    data = query.data

    if data.startswith('toggle:'):
        idx = int(data.split(':', 1)[1])
        if idx in selected:
            selected.discard(idx)
        else:
            selected.add(idx)
        await query.answer()
        await query.edit_message_reply_markup(reply_markup=_gabung_keyboard(chat_id))
        return

    if data.startswith('remove:'):
        idx = int(data.split(':', 1)[1])
        if 0 <= idx < len(entries):
            removed = entries.pop(idx)
            # geser index yang lebih besar dari idx supaya tetap nyambung
            # ke entri yang sama setelah satu dihapus dari tengah daftar
            new_selected = set()
            for s in selected:
                if s < idx:
                    new_selected.add(s)
                elif s > idx:
                    new_selected.add(s - 1)
            SELECTIONS[chat_id] = new_selected
            await query.answer(f'"{removed["label"]}" dihapus dari sesi.')
        else:
            await query.answer()
        if not entries:
            await query.edit_message_text('Sesi sekarang kosong. Upload rekening baru lalu ketik /gabung lagi.')
            return
        await query.edit_message_reply_markup(reply_markup=_gabung_keyboard(chat_id))
        return

    if data == 'reset_all':
        SESSIONS[chat_id] = []
        SELECTIONS[chat_id] = set()
        await query.answer('Sesi dikosongkan.')
        await query.edit_message_text('Semua rekening di sesi ini sudah dihapus. Upload ulang lalu ketik /gabung kalau perlu.')
        return

    if data == 'select_all':
        SELECTIONS[chat_id] = set(range(len(entries)))
        await query.answer()
        await query.edit_message_reply_markup(reply_markup=_gabung_keyboard(chat_id))
        return

    if data == 'select_none':
        SELECTIONS[chat_id] = set()
        await query.answer()
        await query.edit_message_reply_markup(reply_markup=_gabung_keyboard(chat_id))
        return

    if data == 'cancel':
        await query.answer('Dibatalkan.')
        await query.edit_message_text('Digagalkan — sesi rekeningmu masih tersimpan, ketik /gabung lagi kapan saja.')
        return

    if data == 'process':
        if not selected:
            await query.answer('Pilih minimal satu rekening dulu.', show_alert=True)
            return
        await query.answer('Memproses...')
        chosen = [entries[i] for i in sorted(selected)]

        bulan_tahun_counter = Counter(
            (e['bulan'], e['tahun']) for e in chosen if e.get('bulan') and e.get('tahun')
        )
        if bulan_tahun_counter:
            (bulan, tahun), _ = bulan_tahun_counter.most_common(1)[0]
            recap_name = f'Rekap {bulan} {tahun}.xlsx'
        else:
            recap_name = 'Rekap Gabungan.xlsx'

        with tempfile.TemporaryDirectory() as tmp:
            out_path = os.path.join(tmp, recap_name)
            recon_entries = [{
                'sheet_title': e['label'], 'rows': e['rows'],
                'saldo_awal': e['saldo_awal'], 'saldo_akhir': e['saldo_akhir'],
            } for e in chosen]
            write_recon_xlsx(recon_entries, out_path)

            # sesi otomatis dikosongkan begitu selesai gabung, supaya tidak
            # kebawa/kegabung lagi tanpa sengaja di /gabung berikutnya
            SESSIONS[chat_id] = []
            SELECTIONS[chat_id] = set()

            await query.edit_message_text(
                f'Digabung {len(chosen)} rekening: ' + ', '.join(e['label'] for e in chosen) +
                '\n\nSesi sudah otomatis dikosongkan.'
            )
            with open(out_path, 'rb') as f:
                await context.bot.send_document(chat_id=chat_id, document=f, filename=os.path.basename(out_path))
        return


def main():
    if not BOT_TOKEN:
        raise SystemExit('BOT_TOKEN env var belum diset.')

    app = (
        Application.builder()
        .token(BOT_TOKEN)
        .read_timeout(60)
        .write_timeout(60)
        .connect_timeout(30)
        .get_updates_read_timeout(60)
        .build()
    )
    app.add_handler(CommandHandler('start', start))
    app.add_handler(CommandHandler('gabung', gabung_command))
    app.add_handler(CommandHandler('sesi', sesi_command))
    app.add_handler(CommandHandler('reset', reset_command))
    app.add_handler(CallbackQueryHandler(gabung_callback))
    app.add_handler(MessageHandler(filters.Document.PDF | filters.Document.FileExtension('xlsx'), handle_document))

    logger.info('Bot jalan...')
    app.run_polling()


if __name__ == '__main__':
    main()
