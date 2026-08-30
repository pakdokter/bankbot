import os
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

HEADERS = [
    'Tanggal', 'Keterangan Transaksi', 'Kategori Transaksi', 'Debit', 'Kredit',
    'Saldo Kumulatif', 'Subjek Transaksi', 'Objek Transaksi', 'Keterangan Tambahan',
]

COL_WIDTHS = [12, 24, 24, 16, 16, 16, 20, 24, 40]

# --- styling: Arial 9 everywhere, Stoa green branding, light highlights ----
FONT_NAME = 'Arial'
FONT_SIZE = 9

FONT_DEFAULT = Font(name=FONT_NAME, size=FONT_SIZE)
FONT_BOLD = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
FONT_HEADER = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color='FFFFFF')

HEADER_FILL = PatternFill(start_color='1F5C22', end_color='1F5C22', fill_type='solid')
ANCHOR_FILL = PatternFill(start_color='FDF0C7', end_color='FDF0C7', fill_type='solid')   # Saldo Awal/Akhir
SUMMARY_FILL = PatternFill(start_color='D9EAD9', end_color='D9EAD9', fill_type='solid')  # ringkasan bawah
ALT_ROW_FILL = PatternFill(start_color='EAF4EA', end_color='EAF4EA', fill_type='solid')  # baris genap

THIN_GREEN = Side(style='thin', color='A9C9AC')
CELL_BORDER = Border(left=THIN_GREEN, right=THIN_GREEN, top=THIN_GREEN, bottom=THIN_GREEN)

INDO_MONTHS = [
    'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
    'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember',
]


def month_name(month_num):
    """month_num: int or numeric string (1-12) -> Indonesian month name."""
    try:
        return INDO_MONTHS[int(month_num) - 1]
    except (TypeError, ValueError, IndexError):
        return ''


def sanitize_sheet_title(title):
    """Excel sheet names forbid \\ / ? * [ ] : and can't be empty or >31 chars."""
    title = re.sub(r'[\\/:*?\[\]]', '-', title or '')
    title = title.strip() or 'Sheet'
    return title[:31]


def sanitize_filename(name, ext='xlsx'):
    """Sanitizes a string for use as a filename (not a sheet title -- Excel
    sheet names and OS filenames forbid different character sets)."""
    name = re.sub(r'[\\/:*?"<>|]', '_', (name or 'Sheet').strip())
    return f'{name}.{ext}'


def build_filename(self_code, bulan, tahun, ext='xlsx'):
    """'BCA-887', 'Januari', '2025' -> 'BCA-887 Januari 2025.xlsx'"""
    parts = [p for p in (self_code or 'Rekening', bulan or '', str(tahun or '')) if p]
    name = ' '.join(parts).strip()
    return sanitize_filename(name, ext)


def sheet_title_from_meta(meta):
    parts = [p for p in (meta.get('self_code') or 'Rekening', meta.get('bulan') or '', str(meta.get('tahun') or '')) if p]
    return sanitize_sheet_title(' '.join(parts).strip())

# Konvensi tanda: Debit = uang KELUAR (disimpan NEGATIF, tampil dalam kurung
# lewat number_format akuntansi), Kredit = uang MASUK (positif). Ini mengikuti
# label DB/CR asli tiap bank, bukan konvensi buku besar aset (yang membalik
# Debit/Kredit). Kalau ternyata maunya dibalik, gampang ditukar di satu
# tempat ini (ROW_BUILDER di bawah).
DEBIT_NUMBER_FORMAT = '"Rp" #,##0.00;("Rp" #,##0.00)'
KREDIT_NUMBER_FORMAT = '"Rp" #,##0.00'
SALDO_NUMBER_FORMAT = '"Rp" #,##0.00'

# --- kode singkat rekening Stoa sendiri, untuk kolom Subjek/Objek saat ------
# transaksinya adalah pemindahan uang ANTAR rekening milik Stoa sendiri
# (bukan ke pihak luar) -- memudahkan rekonsiliasi lintas rekening.
ACCOUNT_CODES = {
    'bri_personal': 'BRI-507',
    'bri_business': 'BRI-567(Biz)',
    'bca_887': 'BCA-887',
    'bca_417': 'BCA-417',
    'bca_giro': 'BCA-292(Biz)',
    'jago': 'Jago',
}
FLIPTECH_LABEL = 'Rekening Lain (via Fliptech)'

# --- kategori transaksi untuk kebutuhan akuntansi operasional Stoa -------
# Heuristik berbasis kata kunci pada objek/catatan, plus lookup merchant dan
# nomor rekening spesifik yang sudah dikonfirmasi user lewat feedback.
# Tetap belum 100% akurat untuk semua kasus (lihat catatan di bawah).

OWNER_KEYWORDS = ('AHMAD ROZIYAN', 'ROZIYAN HIDAYAT', 'OJAN', 'OWNER')
DEBT_KEYWORDS = ('CICILAN', 'ANGSURAN', 'BAYAR SB', 'PINJAM', 'UTANG')
UTILITY_KEYWORDS = ('SEWA', 'LISTRIK', ' PLN', 'AIR STO', 'UTILITAS')
WALLET_KEYWORDS = ('SHOPEE', 'OVO ', ' OVO', 'GOPAY', 'DANA ', 'TELKOMSEL', 'TOP UP', 'ISI SALDO', 'PULSA', 'BRIVA')
PURCHASE_KEYWORDS = ('BELANJA', 'BELI ', 'ONGKIR', 'SUPPLIER', 'GANTI UANG BELANJA')
PAYROLL_KEYWORDS = ('GAJI',)
TRANSFER_TYPES = ('TRANSFER', 'TRSF', 'BI-FAST', 'SWITCHING', 'KIRIM')

MERCHANT_CATEGORY_MAP = {
    'CV HARAPAN KITA': 'Belanja Operasional',
    'CV HARAPAN': 'Belanja Operasional',
    'MIRA LAUNDRY': 'Belanja Operasional',
    'ORIEL CHICKEN': 'Belanja Konsumsi',
    'DECO COFFEE': 'Belanja Konsumsi',
    'TRANSFERPAY': 'Penjualan',
    'NIDAUL JIHAD': 'Belanja Bahan',
    'DEWA AYU EKA FERR': 'Belanja Konsumsi',
    'ADOBE': 'Belanja Operasional',
    'VISIONET': 'Penjualan',
}

KNOWN_OWNER_ACCOUNTS = {
    '473501000343538',
}

# entitas/subjek pengirim yang selalu berarti modal masuk pemilik, walau
# bukan si pemilik sendiri (mis. rekening perusahaan/keluarga yang dipakai
# untuk suntik modal secara rutin)
KNOWN_MODAL_ENTITIES = (
    'BUANA MEDIA TEKNOL',
)

# Alias pegawai/pemilik yang sudah dikonfirmasi -- dipakai bareng oleh semua
# parser yang perlu mendeteksi pola "Gaji <Nama> [Bulan]". Objek transaksi
# selalu ditulis dengan nama kanonik di sini, apa pun varian yang muncul.
EMPLOYEE_ALIASES = {
    'LATIFATUL HUSNA': 'Latifatul Husna',
    'EVA': 'Latifatul Husna',
    'ROZIYAN HIDAYAT': 'Ahmad Roziyan Hidayat',
    'OJAN': 'Ahmad Roziyan Hidayat',
    'KAK OJAN': 'Ahmad Roziyan Hidayat',
    'OWNER': 'Ahmad Roziyan Hidayat',
    'AHMAD ROZIYAN HIDAYAT': 'Ahmad Roziyan Hidayat',
}
BULAN_PATTERN = 'Januari|Februari|Maret|April|Mei|Juni|Juli|Agustus|September|Oktober|November|Desember'
GAJI_RE = re.compile(rf'^Gaji\s+(.+?)(?:\s+({BULAN_PATTERN})(?:\s+\d{{4}})?)?$', re.I)


def match_gaji(keterangan):
    """Returns (canonical_employee_name, month_text_or_None) if keterangan
    matches the 'Gaji <Name> [Month]' pattern, else None."""
    m = GAJI_RE.match((keterangan or '').strip())
    if not m:
        return None
    name_raw = m.group(1).strip()
    canonical = EMPLOYEE_ALIASES.get(name_raw.upper(), name_raw.title())
    return canonical, (m.group(2).title() if m.group(2) else None)


CATEGORIES_REFERENCE = [
    'Saldo Awal',
    'Penjualan',
    'Belanja Bahan',
    'Belanja Konsumsi',
    'Belanja Operasional',
    'Marketing',
    'Reparasi',
    'Belanja Assets',
    'Gaji Bulan Ini',
    'Riset dan Pengembangan',
    'Tip/Minus/Lebih',
    'Penarikan',
    'Penerimaan',
    'Pembayaran Hutang',
    'Modal & Setoran Pemilik',
    'Pindah Rekening Internal',
    'Transaksi Internal',
    'Biaya Admin & Pajak Bank',
    'Transfer Lainnya',
    'New Kategori',
]

# --- kosakata resmi "bot rekonsiliasi" (reconbot) ---------------------------
# Kategori di sini HARUS persis sama dengan string yang dicek rumus
# SUMIF/SUMIFS di reconbot, supaya baris yang dihasilkan bot konversi ini
# langsung nyambung tanpa perlu di-rename manual saat direkonsiliasi.
# Kategori "Gaji*" pakai wildcard karena reconbot sendiri menerima variasi
# label Gaji (Bulan Ini/Accrual/nama bulan spesifik) via jaring pengaman
# terpisah -- bukan berarti boleh bebas, tapi harus tetap diawali "Gaji ".
RECON_KNOWN_CATEGORIES = {
    'Saldo Awal',
    'Penjualan',
    'Belanja Bahan',
    'Belanja Operasional',
    'Belanja Konsumsi',
    'Marketing',
    'Reparasi',
    'Belanja Assets',
    'Biaya Admin & Pajak Bank',
    'Biaya Admin dan Bunga Bank',
    'Bunga dan Admin Bank',
    'Tip/Minus/Lebih',
    'Penarikan',
    'Penerimaan',
    'Pembayaran Hutang',
    'Pindah Rekening Internal',
    'Pindang Rekening Internal',
    'Transfer Internal',
    'Transfer Lainnya',
    'Transaksi Internal',
    'Modal & Setoran Pemilik',
}


def enforce_recon_category(kategori):
    """Terapkan kosakata kategori resmi reconbot. Kategori "Gaji ..." apa
    pun (dinamis per bulan/pegawai) dianggap dikenal lewat wildcard, sama
    seperti reconbot sendiri. Kategori lain yang tidak persis cocok dengan
    RECON_KNOWN_CATEGORIES ditulis sebagai "New Kategori" supaya kelihatan
    jelas perlu ditambahkan ke reconbot, bukan diam-diam salah kategori."""
    kat = (kategori or '').strip()
    if not kat:
        return 'New Kategori'
    if kat in RECON_KNOWN_CATEGORIES:
        return kat
    if kat.upper().startswith('GAJI'):
        return kat
    return 'New Kategori'

# --- keterangan universal ---------------------------------------------------
# Menormalkan label mentah tiap bank ke kosakata yang sama, supaya baris di
# BCA/BRI/Jago bisa dibandingkan langsung saat rekonsiliasi antar rekening.

UNIVERSAL_KETERANGAN_EXACT = {
    'BUNGA': 'Bunga Bank',
    'PAJAK BUNGA': 'Pajak Bunga',
    'BIAYA ADMIN': 'Biaya Admin',
    'BIAYA ADM': 'Biaya Admin',
    'SALDO AWAL': 'Saldo Awal',
    'PENJUALAN QRIS': 'Penjualan QRIS',
    'TRANSAKSI DEBIT': 'Pembayaran QRIS',
    'TRANSAKSI KREDIT': 'Penjualan QRIS',
    'PEMBAYARAN QRIS': 'Pembayaran QRIS',
    'QRIS': 'Pembayaran QRIS',
}


def normalize_keterangan(keterangan, debit, kredit):
    ket = (keterangan or '').strip()
    ket_upper = ket.upper()
    if ket_upper in UNIVERSAL_KETERANGAN_EXACT:
        return UNIVERSAL_KETERANGAN_EXACT[ket_upper]
    if 'TELKOMSEL' in ket_upper:
        return 'Pulsa/Kuota'
    if 'FLIPTECH' in ket_upper:
        return 'Pindah Rekening Internal via Fliptech'
    if ' TO ' in ket_upper or ket_upper.startswith(('TRANSFER', 'TRSF', 'BI-FAST', 'SWITCHING')):
        return 'Transfer Masuk' if kredit else 'Transfer Keluar'
    return ket


def _merchant_match(objek):
    ob = (objek or '').upper()
    for merchant, cat in MERCHANT_CATEGORY_MAP.items():
        if merchant in ob:
            return cat
    return None


def categorize(keterangan, objek, catatan, debit, kredit):
    """keterangan is expected to already be the *normalized* universal label
    (see normalize_keterangan) — categorize() is always called after that
    step in write_xlsx. Return value is always run through
    enforce_recon_category() before this function returns, so callers never
    need to double-check the result against RECON_KNOWN_CATEGORIES."""
    ket = (keterangan or '').upper()
    ob = (objek or '').upper()
    text = f"{keterangan or ''} {objek or ''} {catatan or ''}".upper()

    kat = _categorize_raw(ket, ob, text, debit, kredit)
    return enforce_recon_category(kat)


def _categorize_raw(ket, ob, text, debit, kredit):
    if ket == 'SALDO AWAL':
        return 'Saldo Awal'
    if ket in ('BUNGA', 'BUNGA BANK', 'PAJAK BUNGA', 'BIAYA ADMIN', 'BIAYA ADM', 'ADMIN TRANSFER'):
        return 'Biaya Admin & Pajak Bank'
    if 'BIAYA PEMBAYARAN' in text:
        return 'Biaya Admin & Pajak Bank'
    if 'FLIPTECH' in text:
        return 'Transaksi Internal'
    if ' TO ' in text or 'IBIZ' in text or 'NBMB' in text:
        return 'Pindah Rekening Internal'

    if any(k in ob for k in KNOWN_MODAL_ENTITIES):
        return 'Modal & Setoran Pemilik'

    merchant_hit = _merchant_match(ob)
    if merchant_hit:
        return merchant_hit

    if ket == 'PENJUALAN QRIS':
        return 'Penjualan'
    if kredit and 'QRIS' in text:
        return 'Penjualan'

    if any(k in text for k in PAYROLL_KEYWORDS):
        return 'Gaji Bulan Ini'

    if any(acc in ob for acc in KNOWN_OWNER_ACCOUNTS):
        return 'Modal & Setoran Pemilik'
    if any(k in ob for k in OWNER_KEYWORDS):
        return 'Modal & Setoran Pemilik'
    if 'TARIK TUNAI' in text or ('SETOR' in text and 'SETORAN' not in ket):
        return 'Modal & Setoran Pemilik'
    if 'TARIKAN' in text and 'ATM' in text:
        return 'Belanja Operasional'

    if any(k in text for k in DEBT_KEYWORDS):
        return 'Pembayaran Hutang'
    if any(k in text for k in UTILITY_KEYWORDS):
        return 'Belanja Operasional'
    if any(k in text for k in WALLET_KEYWORDS):
        return 'Belanja Operasional'
    if any(k in text for k in PURCHASE_KEYWORDS):
        return 'Belanja Operasional'
    if any(k in ket for k in TRANSFER_TYPES):
        return 'Belanja Operasional' if debit else 'Transfer Lainnya'
    # kategori generik yang belum kena aturan spesifik apa pun -- selama
    # uangnya keluar, anggap belanja operasional biasa dulu daripada
    # dibiarkan tak terkategori terus
    if debit:
        return 'Belanja Operasional'
    return ''


def resolve_party(name, self_code, entity_code_map):
    """Map a raw counterparty name to a short account code if it's one of
    Stoa's own known accounts; otherwise return the name unchanged."""
    if not name:
        return name
    up = name.upper()
    if 'FLIPTECH' in up:
        return FLIPTECH_LABEL
    for needle, code in entity_code_map.items():
        if needle in up:
            return code
    return name


def apply_universal_fields(rows, self_code='', entity_code_map=None):
    """Normalizes keterangan, fills kategori, and derives Subjek/Objek from
    transaction direction + the account holder's own short code. Call this
    once per statement, from each parser's build_rows (or let write_xlsx do
    it lazily with self_code='')."""
    entity_code_map = entity_code_map or {}
    for r in rows:
        debit, kredit = r.get('debit'), r.get('kredit')
        raw_ket = r.get('keterangan')
        # categorize() looks for bank-specific raw markers (" TO ", "IBIZ",
        # "FLIPTECH", ...) that normalize_keterangan() already collapses
        # away, so category detection must run on the RAW label.
        r['kategori'] = categorize(raw_ket, r.get('objek'), r.get('catatan'), debit, kredit)
        r['keterangan'] = normalize_keterangan(raw_ket, debit, kredit)

        counterparty = resolve_party(r.get('objek', ''), self_code, entity_code_map)
        if not counterparty or counterparty == '-':
            if r['kategori'].strip().lower().startswith('belanja'):
                counterparty = 'Tenant Lain'
        if r.get('_is_opening_balance'):
            r['subjek'], r['objek'] = '-', '-'
        elif debit:
            r['subjek'], r['objek'] = (self_code or 'Rekening Ini'), counterparty
        elif kredit:
            r['subjek'], r['objek'] = counterparty, (self_code or 'Rekening Ini')
        else:
            r['subjek'], r['objek'] = '', counterparty
    return rows


def _ensure_no_blank_fields(r, self_code=''):
    """Jaring pengaman terakhir: Keterangan/Kategori/Subjek/Objek tidak
    boleh pernah kosong di output, apa pun jalur pemrosesannya sebelumnya."""
    if not (r.get('keterangan') or '').strip():
        r['keterangan'] = 'Transaksi Tanpa Keterangan'
    r['kategori'] = enforce_recon_category(r.get('kategori'))
    if not (r.get('subjek') or '').strip():
        r['subjek'] = self_code or '-'
    if not (r.get('objek') or '').strip():
        r['objek'] = self_code or '-'


def populate_sheet(ws, rows, self_code='', entity_code_map=None, saldo_awal=None, saldo_akhir=None):
    """Fills one worksheet with the standard 9-column layout: header, an
    optional opening-balance row, all transaction rows, and a closing
    summary block. Shared by write_xlsx (single account) and
    write_recon_xlsx (multiple accounts, one sheet each)."""
    if rows and 'subjek' not in rows[0]:
        apply_universal_fields(rows, self_code, entity_code_map)

    # jaring pengaman terakhir yang berlaku ke SEMUA jalur (termasuk
    # kasir.py/preformatted.py yang mengisi subjek/objek/kategori sendiri
    # dan tidak lewat apply_universal_fields di atas) -- Keterangan,
    # Kategori, Subjek, Objek tidak boleh pernah kosong, dan Kategori
    # selalu ditulis pakai kosakata resmi reconbot (atau "New Kategori").
    for r in rows:
        _ensure_no_blank_fields(r, self_code)

    ws.append(HEADERS)
    for c in ws[1]:
        c.font = FONT_HEADER
        c.alignment = Alignment(horizontal='center')
        c.fill = HEADER_FILL
        c.border = CELL_BORDER
    ws.freeze_panes = 'A2'

    anchor_rows = []
    if saldo_awal is not None:
        ws.append(['', 'Saldo Awal', 'Saldo Awal Bulan', None, saldo_awal, saldo_awal, '-', '-', None])
        anchor_rows.append(ws.max_row)

    total_debit, total_kredit = 0.0, 0.0
    data_start_row = ws.max_row + 1
    for r in rows:
        debit = r.get('debit')
        kredit = r.get('kredit')
        if debit:
            total_debit += debit
        if kredit:
            total_kredit += kredit
        ws.append([
            r['tanggal'], r['keterangan'], r.get('kategori', ''),
            debit, kredit, r.get('saldo'),
            r.get('subjek', ''), r.get('objek', ''), r.get('catatan', ''),
        ])

    data_last_row = ws.max_row

    # baris data: font seragam + border rapi + selang-seling warna baris genap
    for row in ws.iter_rows(min_row=data_start_row, max_row=data_last_row):
        is_alt = (row[0].row - data_start_row) % 2 == 1
        for cell in row:
            cell.font = FONT_DEFAULT
            cell.border = CELL_BORDER
            if is_alt:
                cell.fill = ALT_ROW_FILL

    for r_idx in anchor_rows:
        for cell in ws[r_idx]:
            cell.font = FONT_BOLD
            cell.fill = ANCHOR_FILL
            cell.border = CELL_BORDER

    ws.append([])
    summary_start = ws.max_row + 1
    ws.append(['', 'Saldo Awal', '', None, None, saldo_awal, '', '', ''])
    ws.append(['', 'Total Debit (Uang Keluar)', '', total_debit or None, None, None, '', '', ''])
    ws.append(['', 'Total Kredit (Uang Masuk)', '', None, total_kredit or None, None, '', '', ''])
    final_saldo = saldo_akhir if saldo_akhir is not None else (rows[-1]['saldo'] if rows else saldo_awal)
    ws.append(['', 'Saldo Akhir', '', None, None, final_saldo, '', '', ''])
    for row in ws.iter_rows(min_row=summary_start, max_row=ws.max_row):
        for cell in row:
            cell.font = FONT_BOLD
            cell.fill = SUMMARY_FILL
            cell.border = CELL_BORDER

    for row in ws.iter_rows(min_row=2, max_row=data_last_row, min_col=4, max_col=4):
        for cell in row:
            if cell.value is not None:
                cell.number_format = DEBIT_NUMBER_FORMAT
    for row in ws.iter_rows(min_row=2, max_row=data_last_row, min_col=5, max_col=5):
        for cell in row:
            if cell.value is not None:
                cell.number_format = KREDIT_NUMBER_FORMAT
    for row in ws.iter_rows(min_row=2, max_row=data_last_row, min_col=6, max_col=6):
        for cell in row:
            if cell.value is not None:
                cell.number_format = SALDO_NUMBER_FORMAT
    for row in ws.iter_rows(min_row=summary_start, max_row=ws.max_row, min_col=4, max_col=6):
        for cell in row:
            if cell.value is not None:
                cell.number_format = DEBIT_NUMBER_FORMAT if cell.column_letter == 'D' else (
                    KREDIT_NUMBER_FORMAT if cell.column_letter == 'E' else SALDO_NUMBER_FORMAT)

    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.row_dimensions[1].height = 18


def write_xlsx(rows, out_path, sheet_title='Mutasi', self_code='', entity_code_map=None,
               saldo_awal=None, saldo_akhir=None):
    """rows: list of dicts with keys tanggal, keterangan, debit (negative or
    None), kredit (positive or None), saldo, objek, catatan. This function:
    normalizes keterangan/kategori/subjek/objek uniformly, prepends a Saldo
    Awal row (if saldo_awal is given), and appends a closing summary block
    (Saldo Awal / Total Debit / Total Kredit / Saldo Akhir)."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    populate_sheet(ws, rows, self_code, entity_code_map, saldo_awal, saldo_akhir)
    wb.save(out_path)
    return out_path


def write_recon_xlsx(entries, out_path, include_blank_recon_tab=True):
    """entries: list of dicts, one per account/kantong, each already fully
    processed (rows carry subjek/objek/kategori already), with keys:
    sheet_title, rows, saldo_awal, saldo_akhir. Writes one combined workbook
    with one sheet per entry (each in the same 9-column layout), plus an
    empty 'Rekonsiliasi' tab at the end as a starting point for manual
    cross-referencing."""
    wb = Workbook()
    wb.remove(wb.active)
    used_titles = set()
    for e in entries:
        title = sanitize_sheet_title(e['sheet_title'] or 'Sheet')
        base, n = title, 2
        while title in used_titles:
            suffix = f' ({n})'
            title = base[:31 - len(suffix)] + suffix
            n += 1
        used_titles.add(title)
        ws = wb.create_sheet(title=title)
        populate_sheet(ws, e['rows'], saldo_awal=e.get('saldo_awal'), saldo_akhir=e.get('saldo_akhir'))
    if include_blank_recon_tab:
        wb.create_sheet(title='Rekonsiliasi')
    wb.save(out_path)
    return out_path


def split_workbook(xlsx_path, out_dir):
    """Splits every sheet of an .xlsx into its own standalone file, named
    after the sheet (formatting/styles preserved as-is). Returns a list of
    (sheet_name, output_path). Used by the bot's /pisah-style flow -- the
    reverse of write_recon_xlsx."""
    probe = load_workbook(xlsx_path, read_only=True)
    sheet_names = list(probe.sheetnames)
    probe.close()

    outputs = []
    used_names = set()
    for name in sheet_names:
        wb = load_workbook(xlsx_path)
        for other in list(wb.sheetnames):
            if other != name:
                del wb[other]

        fname = sanitize_filename(name)
        base, n = fname, 2
        while fname in used_names:
            stem = base.rsplit('.', 1)[0]
            fname = f'{stem} ({n}).xlsx'
            n += 1
        used_names.add(fname)

        out_path = os.path.join(out_dir, fname)
        wb.save(out_path)
        outputs.append((name, out_path))
    return outputs
