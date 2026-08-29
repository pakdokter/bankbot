import os
import re
import sys
import openpyxl
from .common import write_xlsx, month_name, build_filename, match_gaji

MONTH_MAP = {
    'JANUARI': 1, 'FEBRUARI': 2, 'MARET': 3, 'APRIL': 4, 'MEI': 5, 'JUNI': 6,
    'JULI': 7, 'AGUSTUS': 8, 'SEPTEMBER': 9, 'OKTOBER': 10, 'NOVEMBER': 11, 'DESEMBER': 12,
}

# Nama toko/tenant yang sudah dikenal (dari data kasir yang berulang dan
# cross-reference dengan merchant yang muncul di statement bank Stoa).
# Cocok sebagai prefix ATAU substring, tidak case-sensitive. Kalau ada toko
# baru yang sering muncul tapi belum masuk sini, tambahkan ke daftar ini --
# ini bukan daftar resmi dari stoabot, cuma yang bisa dikenali dari data.
KNOWN_TOKO = {
    'ALFAMART': 'Alfamart',
    'MAK OPIK': 'Mak Opik',
    'MAH OPIK': 'Mak Opik',
    'MIRA': 'Mira Laundry',
    'SB': 'SB (Sinar Bahagia)',
    'PASAR': 'Pasar',
    'PRIMER': 'Primer',
    'DINDA FOOD': 'Dinda Food',
    'TOKO YOGA': 'Toko Yoga',
    'MADAM': 'Madam Baha',
    'TOKO SURYA': 'Toko Surya',
    'WAROENG': 'Waroeng',
    'CV HARAPAN': 'CV Harapan Kita',
    'ORIEL CHICKEN': 'Oriel Chicken',
    'DECO COFFEE': 'Deco Coffee',
    'NIDAUL JIHAD': 'Nidaul Jihad',
}
TENANT_LAIN = 'Tenant Lain'

OWNER_NAME_MARKERS = ('OJAN', 'IYAN', 'ROZIYAN')
GENERIC_MONEY_WORDS = {'LEBIH', 'KURANG', 'MINUS', 'KEMBALI', 'CUSTOMER'}

SETORAN_RE = re.compile(r'^Setoran\s+ke\s+(.+)$', re.I)
SETORAN_TUNAI_RE = re.compile(r'SETORAN\s*TUNAI', re.I)


def match_toko(text):
    up = (text or '').upper().strip()
    for needle, canonical in KNOWN_TOKO.items():
        if up.startswith(needle) or needle in up:
            return canonical
    return None


def split_store_item(desc):
    """Return (store_or_None, item_text) from a kasir 'Keterangan' string."""
    for sep in (' – ', ' - ', '-'):
        if sep in desc:
            left, right = desc.split(sep, 1)
            store = match_toko(left)
            if store:
                return store, right.strip()
            return None, desc.strip()
    store = match_toko(desc)
    if store:
        # strip the matched prefix off, if there's more text after it
        up = desc.upper()
        for needle in KNOWN_TOKO:
            if up.startswith(needle):
                rest = desc[len(needle):].strip(' -–')
                return store, (rest if rest else desc.strip())
        return store, desc.strip()
    return None, desc.strip()


def categorize_kasir(kategori_kasir, item_text, person_name=None):
    it = (item_text or '').upper()
    if kategori_kasir == 'Penjualan':
        return 'Penjualan'
    if kategori_kasir == 'Penarikan':
        if person_name and any(m in person_name.upper() for m in OWNER_NAME_MARKERS):
            return 'Modal & Setoran Pemilik'
        return 'Gaji & Tenaga Kerja'
    if kategori_kasir in ('Penerimaan', 'Koreksi'):
        return 'Lainnya / Perlu Verifikasi'
    if any(k in it for k in ('LAUNDRY', 'PARKIR', 'TISU', 'OJEK', 'ONGKIR', 'LAP ')):
        return 'Belanja Operasional'
    if 'LISTRIK' in it:
        return 'Sewa & Utilitas'
    return 'Belanja Bahan'


HEADER_ALIASES = {
    'TANGGAL': 'tanggal',
    'KETERANGAN': 'keterangan', 'KETERANGAN / DESKRIPSI': 'keterangan',
    'KATEGORI': 'kategori_kasir', 'KATEGORI TRANSAKSI': 'kategori_kasir',
    'DEBIT': 'debit', 'DEBIT (RP)': 'debit',
    'KREDIT': 'kredit', 'KREDIT (RP)': 'kredit',
    'SALDO': 'saldo', 'SALDO (RP)': 'saldo', 'SALDO KUMULATIF': 'saldo',
    'SESI': 'sesi',
    'FLAG': 'flag', 'KETERANGAN TAMBAHAN': 'flag',
    'BULAN': 'bulan_text',
    'BULAN (ANGKA)': 'bulan_angka',
}


def find_header_row(ws, max_scan=5):
    """Return (header_row_index, {field_key: col_index}). Some kasir exports
    have a title on row 1 and the real header on row 2; others put the
    header directly on row 1 — detect whichever it is by looking for a
    'Tanggal'-like first cell."""
    for r in range(1, max_scan + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        cols = {}
        for i, v in enumerate(row):
            key = HEADER_ALIASES.get(str(v or '').strip().upper())
            if key:
                cols[key] = i
        if 'tanggal' in cols and 'keterangan' in cols:
            return r, cols
    return None, {}


def to_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def build_rows(xlsx_path, sheet_name=None):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    title = ws['A1'].value or ''
    ym = re.search(r'(\d{4})', str(title))
    year = int(ym.group(1)) if ym else None

    header_row, cols = find_header_row(ws)
    if header_row is None:
        raise ValueError(
            'Tidak menemukan kolom "Tanggal" + "Keterangan" di 5 baris pertama — '
            'format file kasir ini belum dikenali. Kirim contoh strukturnya biar disesuaikan.'
        )
    if year is None:
        m = re.search(r'(20\d{2})', os.path.basename(xlsx_path))
        year = int(m.group(1)) if m else None

    def get(row, key, default=None):
        idx = cols.get(key)
        return row[idx] if idx is not None and idx < len(row) else default

    rows = []
    saldo_awal = None
    running = None
    matched_count, tenant_lain_count = 0, 0
    last_bulan_num = None

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row is None or all(c is None for c in row):
            continue
        keterangan = get(row, 'keterangan')
        kategori_kasir = get(row, 'kategori_kasir')
        if not keterangan or not kategori_kasir:
            continue
        keterangan = str(keterangan).strip()
        kategori_kasir = str(kategori_kasir).strip()

        tanggal = get(row, 'tanggal')
        bulan_angka = get(row, 'bulan_angka')
        bulan_text = get(row, 'bulan_text')
        sesi = get(row, 'sesi')
        flag = get(row, 'flag')
        debit_k = to_float(get(row, 'debit'))
        kredit_k = to_float(get(row, 'kredit'))
        saldo_k = to_float(get(row, 'saldo'))

        month = None
        if bulan_angka:
            month = int(bulan_angka)
        elif bulan_text:
            month = MONTH_MAP.get(str(bulan_text).strip().upper())
        if month:
            last_bulan_num = month
        month = month or last_bulan_num

        tgl_str = None
        if hasattr(tanggal, 'strftime'):
            tgl_str = tanggal.strftime('%d/%m/%Y')
        elif tanggal and month and year:
            try:
                tgl_str = f'{int(tanggal):02d}/{month:02d}/{year}'
            except (TypeError, ValueError):
                tgl_str = None


        if kategori_kasir == 'Saldo Awal':
            saldo_awal = debit_k if debit_k is not None else saldo_k
            running = saldo_awal
            continue
        if kategori_kasir == 'Saldo':
            # session checkpoint / signature line, not a real transaction
            continue

        # kasir sheet uses the asset-ledger convention (Debit = uang masuk,
        # Kredit = uang keluar) -- flip to this bot's bank-statement
        # convention (Debit = keluar/negatif, Kredit = masuk/positif)
        my_debit = -kredit_k if kredit_k else None
        my_kredit = debit_k if debit_k else None

        person_name = None
        if kategori_kasir == 'Penarikan':
            m = re.search(r'Tarik Tunai\s+(.+)', keterangan, re.I)
            person_name = m.group(1).strip() if m else None
        elif kategori_kasir == 'Penerimaan':
            m = re.search(r'Uang\s+(.+)', keterangan, re.I)
            if m and m.group(1).strip().upper() not in GENERIC_MONEY_WORDS:
                person_name = m.group(1).strip()

        store, item_text = split_store_item(keterangan)
        if store:
            matched_count += 1
            toko = store
        else:
            toko = TENANT_LAIN
            if kategori_kasir == 'Pengeluaran':
                tenant_lain_count += 1

        kategori = categorize_kasir(kategori_kasir, item_text, person_name)

        # --- pola khusus yang sudah dikonfirmasi lewat feedback: berlaku di
        # semua dokumen kasir, bukan cuma satu bulan tertentu ---
        gaji = match_gaji(keterangan)
        if gaji:
            employee_name, gaji_bulan_text = gaji
            toko = employee_name
            gaji_bulan = gaji_bulan_text or month_name(month)
            kategori = f'Gaji Pegawai {gaji_bulan} {year}' if (gaji_bulan and year) else 'Gaji & Tenaga Kerja'

        # "Setoran ke X" / "Setoran Tunai [CDM] [Bank]" -- pemindahan uang
        # dari kas kasir ke rekening bank = transfer antar kantong, bukan
        # pengeluaran/penerimaan biasa.
        m_setor = SETORAN_RE.match(keterangan.strip())
        if m_setor:
            toko = m_setor.group(1).strip()
            kategori = 'Transaksi Internal'
        elif SETORAN_TUNAI_RE.search(keterangan):
            kategori = 'Transaksi Internal'
            m2 = re.search(r'SETORAN\s*TUNAI\s*(?:CDM)?\s*(.*)$', keterangan, re.I)
            extra = m2.group(1).strip() if m2 else ''
            toko = extra if extra else 'Bank'

        if kategori_kasir == 'Penjualan':
            item_text = 'Penjualan'

        if my_debit is not None:
            subjek_field, objek_field = 'Kasir', toko
        else:
            if person_name:
                src = person_name
            elif kategori_kasir == 'Penjualan':
                src = 'Penjualan'
            elif kategori_kasir in ('Penerimaan', 'Koreksi'):
                src = '-'
            else:
                src = toko
            subjek_field, objek_field = src, 'Kasir'

        catatan_parts = []
        if flag:
            catatan_parts.append(str(flag))
        if sesi:
            catatan_parts.append(f'Sesi: {sesi}')

        if running is not None:
            running = round(running + (my_kredit or 0) + (my_debit or 0), 2)
        elif saldo_k is not None:
            running = saldo_k

        rows.append({
            'tanggal': tgl_str,
            'keterangan': item_text or keterangan,
            'debit': my_debit,
            'kredit': my_kredit,
            'saldo': running if running is not None else saldo_k,
            'subjek': subjek_field,
            'objek': objek_field,
            'catatan': '; '.join(catatan_parts),
            'kategori': kategori,
        })

    saldo_akhir = running

    info = {
        'toko_dikenali': matched_count,
        'tenant_lain': tenant_lain_count,
    }
    bulan, tahun = '', ''
    if rows and rows[0]['tanggal']:
        d, m, y = rows[0]['tanggal'].split('/')
        bulan, tahun = month_name(m), y
    meta = {'self_code': 'Kasir', 'bulan': bulan, 'tahun': tahun}
    return rows, saldo_awal, saldo_akhir, info, meta


if __name__ == '__main__':
    xlsx_path = sys.argv[1]
    out_path = sys.argv[2]
    rows, saldo_awal, saldo_akhir, info, meta = build_rows(xlsx_path)
    write_xlsx(rows, out_path, saldo_awal=saldo_awal, saldo_akhir=saldo_akhir)
    print(f'Total baris: {len(rows)}')
    print(f'Toko dikenali: {info["toko_dikenali"]}, Tenant Lain (pengeluaran): {info["tenant_lain"]}')
    print(f'Saldo awal: {saldo_awal}, Saldo akhir: {saldo_akhir}')
    print('Nama file disarankan:', build_filename(meta['self_code'], meta['bulan'], meta['tahun']))
