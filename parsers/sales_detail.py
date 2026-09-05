"""Parser for POS "Detail Penjualan" sales exports (kolom No Transaksi,
Metode Pembayaran, Status Pembayaran, dst).

Laporan ini BUKAN rekening/kasir sungguhan -- ini interpretasi: tiap
transaksi dipetakan ke KEMANA uangnya akan mendarat (kas fisik, atau
rekening bank tempat QRIS/kartu settle), supaya bisa dicocokkan manual
sama statement bank/kasir yang sebenarnya nanti (lewat /gabung).

Satu file bisa menghasilkan beberapa "kelompok tujuan" sekaligus (Cash,
QRIS+Kartu BRI, QRIS+Kartu BCA, dst) -- masing-masing jadi satu sheet
sendiri di output, formatnya sama seperti account biasa (9 kolom + Saldo
Awal + ringkasan), meski "Saldo"-nya di sini cuma total kumulatif
penjualan channel itu, bukan saldo rekening beneran.
"""
import re
from collections import Counter
import openpyxl

from .common import month_name

HEADER_NAMES = {
    'NO TRANSAKSI': 'no_transaksi',
    'WAKTU ORDER': 'waktu_order',
    'WAKTU BAYAR': 'waktu_bayar',
    'TOTAL PENJUALAN (RP)': 'total',
    'METODE PEMBAYARAN': 'metode',
    'TIPE PEMBAYARAN': 'tipe',
    'STATUS PEMBAYARAN': 'status',
    'JUMLAH REFUND (RP)': 'jumlah_refund',
}

BULAN_PATTERN = 'Januari|Februari|Maret|April|Mei|Juni|Juli|Agustus|September|Oktober|November|Desember'
PERIODE_RE = re.compile(rf'({BULAN_PATTERN})\s+(20\d{{2}})', re.I)

# metode/tipe pembayaran -> kelompok tujuan uang. Dicek dengan substring
# (case-insensitive) terhadap gabungan teks Metode+Tipe Pembayaran.
DESTINATION_RULES = [
    (r'QRIS\s*BRI', 'QRIS BRI', 'BRI-507'),
    (r'KARTU.*BRI|DEBIT/KREDIT.*BRI', 'QRIS BRI', 'BRI-507'),
    (r'QRIS\s*BCA', 'QRIS BCA', 'BCA-887'),
    (r'KARTU.*BCA|DEBIT/KREDIT.*BCA', 'QRIS BCA', 'BCA-887'),
    (r'\bCASH\b|\bTUNAI\b', 'Cash', 'Kas/Buku'),
]
DESTINATION_RULES = [(re.compile(pat, re.I), grp, code) for pat, grp, code in DESTINATION_RULES]

FALLBACK_GROUP = 'Perlu Verifikasi'
FALLBACK_CODE = 'Perlu Verifikasi'


def _find_header_row(ws, max_scan=20):
    for r in range(1, max_scan + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        cols = {}
        for i, v in enumerate(row):
            key = HEADER_NAMES.get(str(v or '').strip().upper())
            if key:
                cols[key] = i
        if 'no_transaksi' in cols and 'metode' in cols:
            return r, cols
    return None, {}


def is_sales_detail(path, sheet_name=None):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    header_row, _ = _find_header_row(ws)
    return header_row is not None


def _classify(metode, tipe):
    text = f'{metode or ""} {tipe or ""}'
    for pattern, group, code in DESTINATION_RULES:
        if pattern.search(text):
            return group, code
    return FALLBACK_GROUP, FALLBACK_CODE


def _parse_datetime_cell(v):
    """'Waktu Order'/'Waktu Bayar' sudah datetime kalau file-nya asli dari
    export, tapi jaga-jaga kalau berupa teks 'dd-mm-yyyy HH:MM:SS'."""
    if v is None:
        return None
    if hasattr(v, 'strftime'):
        return v
    m = re.match(r'^(\d{2})-(\d{2})-(\d{4})', str(v).strip())
    if m:
        import datetime
        d, mo, y = m.groups()
        return datetime.datetime(int(y), int(mo), int(d))
    return None


def build_groups(path, sheet_name=None):
    """Returns (groups, meta, warnings).
    groups: {group_name: {'rows': [...], 'self_code': str}}
    meta: {'bulan': str, 'tahun': str}
    warnings: list of str (mis. baris dengan metode pembayaran yang belum
    dikenal, masuk ke grup "Perlu Verifikasi")
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    header_row, cols = _find_header_row(ws)
    if header_row is None:
        raise ValueError(
            'Tidak menemukan kolom "No Transaksi" + "Metode Pembayaran" — '
            'format laporan penjualan ini belum dikenali. Kirim contoh strukturnya biar disesuaikan.'
        )

    def get(row, key):
        idx = cols.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    # bulan/tahun dari baris "Periode" di area ringkasan atas (sebelum
    # header_row), fallback ke nama file kalau tidak ketemu
    bulan, tahun = '', ''
    for row in ws.iter_rows(min_row=1, max_row=header_row, values_only=True):
        for v in row:
            if v and isinstance(v, str):
                m = PERIODE_RE.search(v)
                if m:
                    bulan, tahun = m.group(1).title(), m.group(2)
                    break
        if bulan:
            break

    groups = {}
    group_codes = {}
    fallback_methods = Counter()

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row is None or all(c is None for c in row):
            continue
        no_transaksi = get(row, 'no_transaksi')
        if not no_transaksi:
            continue
        status = str(get(row, 'status') or '').strip().upper()
        if status == 'BELUM LUNAS':
            continue  # belum ada uang yang benar-benar bergerak

        metode = get(row, 'metode')
        tipe = get(row, 'tipe')
        group, self_code = _classify(metode, tipe)
        if group == FALLBACK_GROUP:
            fallback_methods[str(metode or '-')] += 1

        waktu = _parse_datetime_cell(get(row, 'waktu_bayar')) or _parse_datetime_cell(get(row, 'waktu_order'))
        tgl_str = waktu.strftime('%d/%m/%Y') if waktu else None

        is_bank_settled = group != 'Cash' and group != FALLBACK_GROUP
        catatan_parts = [f'No Transaksi: {no_transaksi}']
        if is_bank_settled and waktu:
            settle = waktu + __import__('datetime').timedelta(days=1)
            catatan_parts.append(f'Estimasi settle: {settle.strftime("%d/%m/%Y")}')

        if status == 'REFUND':
            jumlah = get(row, 'jumlah_refund') or get(row, 'total') or 0
            keterangan = 'Refund'
            debit, kredit = -abs(float(jumlah)), None
        else:
            jumlah = get(row, 'total') or 0
            keterangan = 'Penjualan'
            debit, kredit = None, abs(float(jumlah))

        entry = groups.setdefault(group, [])
        group_codes[group] = self_code
        entry.append({
            'tanggal': tgl_str,
            'keterangan': keterangan,
            'kategori': 'Penjualan',
            'debit': debit,
            'kredit': kredit,
            'saldo': None,  # dihitung kumulatif di bawah, per grup
            'subjek': 'Penjualan',
            'objek': self_code,
            'catatan': '; '.join(catatan_parts),
        })

    # hitung saldo kumulatif per grup (murni angka penjualan channel itu,
    # BUKAN saldo rekening/kasir beneran -- mulai dari 0, bukan carry-over
    # bulan sebelumnya, karena laporan ini berdiri sendiri per periode)
    result_groups = {}
    for group, rows in groups.items():
        running = 0.0
        for r in rows:
            running = round(running + (r['kredit'] or 0) + (r['debit'] or 0), 2)
            r['saldo'] = running
        result_groups[group] = {'rows': rows, 'self_code': group_codes[group]}

    warnings = []
    if fallback_methods:
        detail = ', '.join(f'{k} ({v}x)' for k, v in fallback_methods.items())
        warnings.append(f'Metode pembayaran belum dikenal, masuk grup "Perlu Verifikasi": {detail}')

    meta = {'bulan': bulan, 'tahun': tahun}
    return result_groups, meta, warnings
