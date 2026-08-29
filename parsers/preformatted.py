"""Parser for XLSX files that are already in this bot's own 9-column output
format (e.g. a manually reconstructed old statement, or a previous bot
output the user edited/annotated). Column order is detected by header name
so extra analysis columns (e.g. a manual "Selisih vs Saldo Tercatat" check)
can sit alongside ours without breaking anything."""
import re
import sys
from collections import Counter
import openpyxl

from .common import HEADERS, write_xlsx, build_filename, month_name, match_gaji

FIELD_KEYS = ['tanggal', 'keterangan', 'kategori', 'debit', 'kredit', 'saldo', 'subjek', 'objek', 'catatan']
HEADER_NAME_MAP = {h.upper(): key for h, key in zip(HEADERS, FIELD_KEYS)}
SELISIH_NAMES = {'SELISIH', 'SELISIH VS SALDO TERCATAT'}

# --- aturan yang sudah dikonfirmasi lewat feedback -- berlaku untuk semua
# dokumen "sudah diolah" berikutnya, bukan cuma satu bulan tertentu ---

# nama yang selalu berarti "Modal Masuk" (uang masuk dari pemilik/keluarga,
# bukan penjualan)
MODAL_MASUK_NAMES = ('AHMAD RIZAN HENDRA',)

# alias merchant/tenant tunggal (nama panjang -> nama pendek kanonik)
TENANT_ALIASES = {
    'PRIMER RAYA': 'Primer',
    'PRIMER': 'Primer',
}
OWNER_MARKERS = ('OWNER', 'ROZIYAN HIDAYAT', 'OJAN', 'KAK OJAN', 'AHMAD ROZIYAN HIDAYAT')
GENERIC_UNRESOLVED_CATEGORIES = {'TRANSFER KELUAR', 'PENGELUARAN', 'TRANSFER LAINNYA'}

# kata kunci -> (kategori, objek_atau_None). Dicek pada gabungan teks
# Keterangan + Keterangan Tambahan, tidak case-sensitive, dengan word
# boundary supaya tidak salah tangkap ("Web" tidak match "Website" dst
# kalau memang perlu lebih ketat -- di sini cukup longgar karena datanya
# manual/singkat).
KEYWORD_RULES = [
    (r'MADAM', 'Belanja Bahan', 'Madam'),
    (r'BEANS', 'Belanja Bahan', None),
    (r'SHOPEE', 'Belanja Bahan', 'Shopee'),
    (r'SINAR BAHAGIA', 'Belanja Bahan', 'Sinar Bahagia'),
    (r'KONSUMSI', 'Belanja Operasional', None),
    (r'\bWEB\b', 'Belanja Operasional', None),
    (r'UTILITIES', 'Belanja Operasional', None),
    (r'SPOTIFY', 'Belanja Operasional', None),
    (r'TELKOM', 'Belanja Operasional', None),
    (r'MR\s*DIY', None, 'MR DIY'),
    (r'PELATIHAN', 'Riset dan Pengembangan', None),
    (r'TARIKAN?\s*ATM', 'Belanja Operasional', None),
    (r'INDOMARET', None, 'Indomaret'),
    (r'MASUYA', None, 'Masuya'),
    (r'ANUGERAH', None, 'Anugerah'),
    (r'AMANAH', 'Belanja Bahan', 'Amanah'),
    (r'APOTEK\s*SURYA\s*FARMA|SURYA\s*FARMA', None, 'Apotek Surya Farma'),
    (r'APOTEK\s*YODAN|\bYODAN\b', None, 'Apotek Yodan'),
    (r'BINTANG\s*PLASTIK', None, 'Bintang Plastik'),
    (r'FADHILAH', 'Belanja Bahan', 'Fadhilah'),
    (r'MAK\s*OPIK|MAH\s*OPIK', 'Belanja Bahan', 'Mak Opik'),
    (r'PASAR\s*PANCOR|\bPASAR\b', 'Belanja Bahan', 'Pasar'),
]
KEYWORD_RULES = [(re.compile(pat, re.I), kat, obj) for pat, kat, obj in KEYWORD_RULES]

KONSUMSI_RE = re.compile(r'KONSUMSI', re.I)
PENJUALAN_RE = re.compile(r'PENJUALAN', re.I)
BUNGA_RE = re.compile(r'BUNGA', re.I)
BIAYA_ADMIN_RE = re.compile(r'BIAYA\s*ADMIN|PAJAK|ADMIN\s*TRANSFER', re.I)
FLIPTECH_RE = re.compile(r'FLIPTECH', re.I)
PRIMER_RE = re.compile(r'PRIMER(?:\s*RAYA)?', re.I)
MODAL_MASUK_KETERANGAN_RE = re.compile(r'^MODAL\s*MASUK$', re.I)
KOREKSI_RE = re.compile(r'KOREKSI', re.I)
# "Setoran Via CDM", "Setoran Tunai", "Setor tunai -> BCA", "Setoran ke BCA", dst
SETORAN_RE = re.compile(r'SETOR(?:AN)?\s*TUNAI|SETORAN', re.I)
BANK_NAME_RE = re.compile(r'\b(BCA|BRI|MANDIRI|JAGO|BSI|BNI|CIMB)\b', re.I)


def _apply_keyword_overrides(keterangan, kategori, objek, catatan, is_kredit=False):
    """Returns (keterangan, kategori, objek) after applying every keyword
    rule confirmed via feedback. Order matters: more specific rules first."""
    text = f'{keterangan} {catatan}'.upper()
    ob_upper = (objek or '').strip().upper()

    if ob_upper in MODAL_MASUK_NAMES or MODAL_MASUK_KETERANGAN_RE.match(keterangan.strip()):
        return 'Modal Masuk', 'Modal & Setoran Pemilik', objek

    if KOREKSI_RE.search(keterangan):
        return 'Tip/Minus', 'Tip/Minus', objek

    gaji = match_gaji(keterangan)
    if gaji:
        employee_name, _bulan_gaji = gaji
        # gaji milik owner yang MASUK (bukan Stoa menggaji dia) sebenarnya
        # penghasilan luar yang disuntikkan sebagai modal, bukan beban gaji
        if is_kredit and any(m in employee_name.upper() for m in OWNER_MARKERS):
            return 'Modal Masuk', 'Modal & Setoran Pemilik', employee_name
        return keterangan, 'Gaji Pegawai', employee_name

    if PENJUALAN_RE.search(kategori) or PENJUALAN_RE.search(keterangan):
        return 'Penjualan', 'Penjualan', objek

    # cek juga kategori ASLI dari sumber -- beberapa file manual sudah
    # menulis "Biaya Admin"/"Pajak Bank"/dst di kolom kategori sendiri
    # walau keterangannya tidak secara harfiah menyebut kata itu
    admin_bunga_text = f'{text} {kategori}'.upper()
    if BIAYA_ADMIN_RE.search(admin_bunga_text):
        return 'Biaya Admin', 'Biaya Admin Bank', objek
    if BUNGA_RE.search(admin_bunga_text):
        return 'Bunga Bank', 'Biaya Admin Bank', objek

    if SETORAN_RE.search(text):
        # setoran tunai kasir <-> bank = perpindahan antar "kantong" Stoa
        # sendiri. Kalau nama bank tujuannya kesebut eksplisit, itu jadi
        # objek (rekening ini yang mengirim); kalau tidak, rekening ini
        # dianggap sisi penerima (bank), objek jadi self_code -- diselesaikan
        # di build_rows setelah self_code diketahui.
        m_bank = BANK_NAME_RE.search(text)
        target = m_bank.group(1).upper() if m_bank else objek
        return keterangan, 'Transaksi Internal', target

    if PRIMER_RE.search(text):
        return 'Belanja Bahan', 'Belanja Bahan', 'Primer'

    new_keterangan = 'Belanja Konsumsi' if KONSUMSI_RE.search(text) else keterangan
    new_kategori, new_objek = kategori, objek
    for pattern, kat, obj in KEYWORD_RULES:
        if pattern.search(text):
            if kat:
                new_kategori = kat
            if obj and (not new_objek or new_objek == '-'):
                new_objek = obj

    # kategori generik/tidak jelas yang belum kena aturan spesifik apa pun di
    # atas -- selama uangnya keluar, anggap sebagai belanja operasional
    # biasa daripada dibiarkan sebagai label transfer mentah
    if new_kategori.strip().upper() in GENERIC_UNRESOLVED_CATEGORIES and not is_kredit:
        new_kategori = 'Belanja Operasional'

    if new_kategori.strip().lower().startswith('belanja') and (not new_objek or new_objek == '-'):
        new_objek = 'Tenant Lain'

    return new_keterangan, new_kategori, new_objek


def _find_columns(ws, max_scan=3):
    for r in range(1, max_scan + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        cols, selisih_col = {}, None
        for i, v in enumerate(row):
            name = str(v or '').strip().upper()
            if name in HEADER_NAME_MAP:
                cols[HEADER_NAME_MAP[name]] = i
            elif name in SELISIH_NAMES:
                selisih_col = i
        if 'tanggal' in cols and 'keterangan' in cols and 'debit' in cols and 'kredit' in cols:
            return r, cols, selisih_col
    return None, {}, None


def is_preformatted(xlsx_path, sheet_name=None):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    header_row, cols, _ = _find_columns(ws)
    return header_row is not None


def to_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def build_rows(xlsx_path, sheet_name=None):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    header_row, cols, selisih_col = _find_columns(ws)
    if header_row is None:
        raise ValueError(
            'File ini bukan format bank yang sudah aku kenali maupun format '
            'output bot sendiri (kolom Tanggal/Keterangan/Debit/Kredit tidak '
            'ketemu). Kirim contoh strukturnya biar disesuaikan.'
        )

    def get(row, key):
        idx = cols.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    # --- pass 1: baca semua baris mentah + tentukan self_code lebih dulu,
    # dari nilai Subjek/Objek asli (sebelum override apa pun) --------------
    raw_rows = []
    party_counter = Counter()
    month_year_counter = Counter()
    saldo_awal, saldo_akhir = None, None
    selisih_flags = 0

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row is None or all(c is None for c in row):
            continue
        keterangan = get(row, 'keterangan')
        if not keterangan:
            continue
        keterangan = str(keterangan).strip()
        if keterangan in ('Total Debit (Uang Keluar)', 'Total Kredit (Uang Masuk)', 'Saldo Akhir'):
            saldo_val = to_float(get(row, 'saldo'))
            if keterangan == 'Saldo Akhir' and saldo_val is not None:
                saldo_akhir = saldo_val
            continue

        tanggal = get(row, 'tanggal')
        kategori = str(get(row, 'kategori') or '').strip()
        debit = to_float(get(row, 'debit'))
        kredit = to_float(get(row, 'kredit'))
        saldo = to_float(get(row, 'saldo'))
        subjek = str(get(row, 'subjek') or '')
        objek = str(get(row, 'objek') or '')
        catatan = str(get(row, 'catatan') or '')

        if kategori.strip().upper().startswith('SALDO AWAL') or keterangan.strip().upper().startswith('SALDO AWAL'):
            # beberapa file manual cuma isi Kredit (atau Debit) untuk baris
            # ini dan biarkan Saldo Kumulatif kosong -- pakai itu sebagai
            # fallback kalau kolom Saldo-nya sendiri tidak terisi
            fallback = kredit if kredit is not None else debit
            saldo_awal = saldo if saldo is not None else (fallback if fallback is not None else saldo_awal)
            continue

        if hasattr(tanggal, 'strftime'):
            tgl_str = tanggal.strftime('%d/%m/%Y')
            month_year_counter[(tanggal.month, tanggal.year)] += 1
        elif isinstance(tanggal, str) and re.match(r'^\d{2}/\d{2}/\d{4}$', tanggal):
            tgl_str = tanggal
            d, m, y = tanggal.split('/')
            month_year_counter[(int(m), int(y))] += 1
        else:
            tgl_str = tanggal if isinstance(tanggal, str) else None

        if selisih_col is not None:
            raw_selisih = row[selisih_col] if selisih_col < len(row) else None
            sv = to_float(raw_selisih)
            if sv and abs(sv) > 0.01:
                selisih_flags += 1

        for p in (subjek, objek):
            p = str(p).strip()
            if p and p != '-':
                party_counter[p] += 1

        raw_rows.append({
            'tanggal': tgl_str, 'keterangan': keterangan, 'kategori': kategori,
            'debit': debit, 'kredit': kredit, 'saldo': saldo,
            'subjek': subjek, 'objek': objek, 'catatan': catatan,
        })

    self_code = party_counter.most_common(1)[0][0] if party_counter else 'Rekening'
    if month_year_counter:
        (m, y), _ = month_year_counter.most_common(1)[0]
        bulan, tahun = month_name(m), y
    else:
        bulan, tahun = '', ''

    # --- pass 2: terapkan aturan kata kunci + pemecahan Fliptech, sekarang
    # self_code sudah diketahui untuk baris Penjualan/Biaya Admin/Bunga Bank
    rows = []
    running = saldo_awal

    def _emit(row_dict):
        """Tambahkan baris ke rows, dan hitung ulang Saldo Kumulatif kalau
        sumbernya tidak mengisi kolom itu (cuma isi Debit/Kredit). Kalau
        sumbernya MEMANG mengisi Saldo, itu dipercaya sebagai checkpoint dan
        running balance disinkronkan ke situ."""
        nonlocal running
        if running is not None:
            delta = (row_dict['kredit'] or 0) + (row_dict['debit'] or 0)
            running = round(running + delta, 2)
            if row_dict['saldo'] is None:
                row_dict['saldo'] = running
            else:
                running = row_dict['saldo']
        rows.append(row_dict)

    for r in raw_rows:
        tgl_str, keterangan, kategori = r['tanggal'], r['keterangan'], r['kategori']
        debit, kredit, saldo = r['debit'], r['kredit'], r['saldo']
        subjek, objek, catatan = r['subjek'], r['objek'], r['catatan']

        if FLIPTECH_RE.search(keterangan) or FLIPTECH_RE.search(catatan):
            total = debit if debit is not None else kredit
            is_debit = debit is not None
            if total is not None:
                abs_total = abs(total)
                main = float((int(abs_total) // 1000) * 1000)
                remainder = round(abs_total - main, 2)
                _emit({
                    'tanggal': tgl_str, 'keterangan': 'Transfer Internal',
                    'kategori': 'Transaksi Internal',
                    'debit': -main if is_debit else None,
                    'kredit': None if is_debit else main,
                    'saldo': None, 'subjek': subjek,
                    'objek': objek or 'Fliptech',
                    'catatan': catatan,
                })
                if remainder:
                    fee_label = 'Biaya Admin' if is_debit else 'Bunga Bank'
                    _emit({
                        'tanggal': tgl_str, 'keterangan': fee_label, 'kategori': 'Biaya Admin Bank',
                        'debit': -remainder if is_debit else None,
                        'kredit': None if is_debit else remainder,
                        'saldo': saldo, 'subjek': '-', 'objek': self_code,
                        'catatan': f'Bagian dari transaksi Fliptech: {keterangan}',
                    })
                continue

        keterangan, kategori, objek = _apply_keyword_overrides(
            keterangan, kategori, objek, catatan, is_kredit=(kredit is not None)
        )

        if kategori == 'Penjualan':
            subjek, objek = 'Penjualan', self_code
        elif kategori == 'Biaya Admin Bank':
            subjek, objek = '-', self_code
        elif kategori == 'Transaksi Internal':
            # dari aturan SETORAN_RE di _apply_keyword_overrides: kalau nama
            # bank tujuan disebut eksplisit di teks, rekening ini yang
            # mengirim (objek = bank tujuan); kalau tidak, rekening ini
            # dianggap sisi bank yang menerima setoran dari kas kasir.
            if objek and objek not in ('-', ''):
                subjek = self_code
            else:
                subjek, objek = 'Kas Kasir', self_code

        _emit({
            'tanggal': tgl_str,
            'keterangan': keterangan,
            'kategori': kategori,
            'debit': debit,
            'kredit': kredit,
            'saldo': saldo,
            'subjek': subjek,
            'objek': objek,
            'catatan': catatan,
        })

    if saldo_akhir is None:
        saldo_akhir = rows[-1]['saldo'] if rows else saldo_awal

    info = {'selisih_flags': selisih_flags}
    meta = {'self_code': self_code, 'bulan': bulan, 'tahun': tahun}
    return rows, saldo_awal, saldo_akhir, info, meta


if __name__ == '__main__':
    xlsx_path = sys.argv[1]
    out_path = sys.argv[2]
    rows, saldo_awal, saldo_akhir, info, meta = build_rows(xlsx_path)
    write_xlsx(rows, out_path, saldo_awal=saldo_awal, saldo_akhir=saldo_akhir)
    print(f'Total baris: {len(rows)}')
    print(f'Saldo awal: {saldo_awal}, Saldo akhir: {saldo_akhir}')
    print('Meta:', meta)
    if info['selisih_flags']:
        print(f'PERINGATAN: {info["selisih_flags"]} baris punya selisih != 0 di kolom Selisih.')
    print('Nama file disarankan:', build_filename(meta['self_code'], meta['bulan'], meta['tahun']))
